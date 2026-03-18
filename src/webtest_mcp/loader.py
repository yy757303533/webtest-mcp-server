"""Excel 用例加载 - 支持自然语言步骤、.xls/.xlsx、单行多步骤"""

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, List, Optional, Union

# 表头别名（支持中英文，宽松匹配）
HEADERS = {
    "case_id": ("case_id", "用例id", "用例ID", "编号", "用例编号"),
    "title": ("title", "标题", "用例名", "用例标题", "用例说明", "用例说明（名称）"),
    "module": ("module", "模块", "功能模块"),
    "test_type": ("test_type", "测试类型", "用例类型"),
    "precondition": ("precondition", "预置条件", "前置条件", "前提条件"),
    "description": ("description", "步骤描述", "操作步骤", "描述", "步骤说明", "测试步骤"),
    "expected": ("expected", "期望结果", "预期结果", "预期"),
    "priority": ("priority", "等级", "优先级"),
    "tags": ("tags", "tags", "标签", "场景标签"),
    # 保留对 DSL 格式的兼容
    "step_no": ("step_no", "步骤号"),
    "action": ("action", "动作", "操作"),
    "target": ("target", "目标", "元素"),
    "value": ("value", "值", "输入值"),
}


@dataclass
class CaseStep:
    """执行步骤（S1. S2. ...）"""
    step_no: int
    description: str


@dataclass
class ExpectedResult:
    """预期结果（E1. E2. ...）"""
    expect_no: int
    description: str


@dataclass
class ExcelCase:
    """
    用例（供 AI 消费）

    执行顺序：先全部执行 steps，再统一验证 expected。
    S 和 E 是独立的两个列表，不是一一对应的。
    """
    case_id: str
    title: str
    steps: List[CaseStep]
    expected: List[ExpectedResult]
    module: str = ""
    test_type: str = ""
    precondition: str = ""
    priority: str = ""
    tags: str = ""


# ---------------------------------------------------------------------------
# Excel 读取（同时支持 .xls 和 .xlsx）
# ---------------------------------------------------------------------------

def _open_workbook(path: Path):
    """
    打开 Excel 文件，返回 (rows, close_fn)。
    - .xls          → xlrd（需安装）
    - .xlsx / .xlsm → openpyxl
    """
    suffix = path.suffix.lower()

    if suffix == ".xls":
        try:
            import xlrd
        except ImportError:
            raise ImportError(
                "读取 .xls 文件需要 xlrd 库，请执行: pip install xlrd"
            )
        wb = xlrd.open_workbook(str(path))
        ws = wb.sheet_by_index(0)
        rows = []
        for r in range(ws.nrows):
            rows.append(tuple(ws.cell_value(r, c) for c in range(ws.ncols)))
        return rows, lambda: None

    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Excel 没有活动工作表")
    rows = list(ws.iter_rows(values_only=True))
    return rows, wb.close


# ---------------------------------------------------------------------------
# 表头匹配
# ---------------------------------------------------------------------------

def _normalize_header(cell_value: Any) -> str:
    """标准化表头：去空白、去换行、转小写"""
    if cell_value is None:
        return ""
    s = str(cell_value).strip().replace("\n", "").replace("\r", "").lower()
    return s


def _find_header_cols(row: tuple) -> dict[str, int]:
    """
    匹配表头列。策略：
    1. 精确匹配（忽略大小写 + 去换行）
    2. 包含匹配（表头包含别名，且长度 < 30）
    """
    cols: dict[str, int] = {}
    alias_exact: dict[str, str] = {}
    alias_contains: list[tuple[str, str]] = []
    for key, aliases in HEADERS.items():
        for a in aliases:
            a_lower = a.lower()
            alias_exact[a_lower] = key
            alias_contains.append((a_lower, key))

    for idx, cell in enumerate(row):
        v = _normalize_header(cell)
        if not v:
            continue
        # 精确匹配优先
        if v in alias_exact:
            key = alias_exact[v]
            if key not in cols:
                cols[key] = idx
            continue
        # 包含匹配（如 "用例编号" 包含 "编号"）
        if len(v) < 30:
            for alias, key in alias_contains:
                if key not in cols and alias in v:
                    cols[key] = idx
                    break
    return cols


def _get_cell(row: tuple, cols: dict[str, int], key: str) -> Any:
    if key not in cols:
        return None
    idx = cols[key]
    if idx >= len(row):
        return None
    return row[idx]


# ---------------------------------------------------------------------------
# 步骤解析（支持单行多步骤：S1. xxx\nS2. xxx）
# ---------------------------------------------------------------------------

_STEP_PATTERN = re.compile(r"[Ss](\d+)[.、]\s*")
_EXPECTED_PATTERN = re.compile(r"[Ee](\d+)[.、]\s*")
_TAG_SPLIT_PATTERN = re.compile(r"[,，;；\s]+")


def _parse_multi_steps(text: str) -> list[tuple[int, str]]:
    """
    解析 "S1. 做xxx\\nS2. 做yyy" 格式，返回 [(step_no, description), ...]
    如果没有 S 编号，整体作为第 1 步返回。
    """
    if not text or not text.strip():
        return []
    parts = _STEP_PATTERN.split(text.strip())
    if len(parts) <= 1:
        return [(1, text.strip())]
    result = []
    i = 1
    while i < len(parts) - 1:
        try:
            step_no = int(parts[i])
        except ValueError:
            step_no = 0
        desc = parts[i + 1].strip()
        if desc:
            result.append((step_no, desc))
        i += 2
    return result if result else [(1, text.strip())]


def _parse_multi_expected(text: str) -> dict[int, str]:
    """
    解析 "E1. 预期xxx\\nE2. 预期yyy" 格式，返回 {1: "预期xxx", 2: "预期yyy"}
    如果没有 E 编号，返回 {0: "整段文本"}（会分配给所有步骤）。
    """
    if not text or not text.strip():
        return {}
    parts = _EXPECTED_PATTERN.split(text.strip())
    if len(parts) <= 1:
        return {0: text.strip()}
    result = {}
    i = 1
    while i < len(parts) - 1:
        try:
            e_no = int(parts[i])
        except ValueError:
            e_no = 0
        exp = parts[i + 1].strip()
        if exp:
            result[e_no] = exp
        i += 2
    return result if result else {0: text.strip()}


# ---------------------------------------------------------------------------
# 主加载函数
# ---------------------------------------------------------------------------

def load_excel_cases(excel_path: Union[str, Path]) -> List[ExcelCase]:
    """
    从 Excel 加载用例，返回自然语言格式。

    支持三种格式：
    1. 单行自然语言：每行一条用例，步骤在一个单元格内（S1. S2.）
    2. 多行自然语言：每行一步，靠 case_id 聚合
    3. DSL 结构化：action + target + value（拼接为自然语言）
    """
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel 文件不存在: {path}")

    rows, close_fn = _open_workbook(path)
    try:
        if not rows:
            raise ValueError("Excel 为空")

        header_cols = _find_header_cols(rows[0])
        has_desc = "description" in header_cols
        has_action = "action" in header_cols
        if not has_desc and not has_action:
            raise ValueError(
                f"Excel 需包含「测试步骤」「步骤描述」或「action」列。"
                f"当前识别到的列: {list(header_cols.keys())}"
            )

        cases_map: dict[str, ExcelCase] = {}

        for row in rows[1:]:
            if not row or all(
                cell is None or (isinstance(cell, str) and not str(cell).strip())
                for cell in row
            ):
                continue

            desc_raw = str(_get_cell(row, header_cols, "description") or "").strip()
            expected_raw = str(_get_cell(row, header_cols, "expected") or "").strip()

            # DSL 格式 fallback
            if not desc_raw and has_action:
                action = str(_get_cell(row, header_cols, "action") or "").strip()
                target = str(_get_cell(row, header_cols, "target") or "").strip()
                value = str(_get_cell(row, header_cols, "value") or "").strip()
                if action:
                    parts = [a for a in [action, target, value] if a]
                    desc_raw = " ".join(parts)

            if not desc_raw:
                continue

            case_id = str(_get_cell(row, header_cols, "case_id") or "").strip() or "default"
            title = str(_get_cell(row, header_cols, "title") or "").strip()
            module = str(_get_cell(row, header_cols, "module") or "").strip()
            test_type = str(_get_cell(row, header_cols, "test_type") or "").strip()
            precondition = str(_get_cell(row, header_cols, "precondition") or "").strip()
            priority = str(_get_cell(row, header_cols, "priority") or "").strip()
            tags = str(_get_cell(row, header_cols, "tags") or "").strip()

            # 解析单行内的多步骤和多预期（S 和 E 是独立列表）
            # 兼容“多行自然语言：每行一步”——若无 S 编号，则按出现顺序自动递增 step_no
            has_step_no = _STEP_PATTERN.search(desc_raw) is not None
            parsed_steps = _parse_multi_steps(desc_raw)
            parsed_expected = _parse_multi_expected(expected_raw)

            if case_id in cases_map:
                existing = cases_map[case_id]
                next_step_no = (max((s.step_no for s in existing.steps), default=0) + 1)
                next_expect_no = (
                    max((e.expect_no for e in existing.expected), default=0) + 1
                )
            else:
                existing = None
                next_step_no = 1
                next_expect_no = 1

            if not has_step_no and len(parsed_steps) == 1 and parsed_steps[0][0] == 1:
                steps = [CaseStep(step_no=next_step_no, description=parsed_steps[0][1])]
            else:
                steps = [CaseStep(step_no=sn, description=sd) for sn, sd in parsed_steps]

            # 兼容“多行自然语言：每行一个预期且无 E 编号”——按出现顺序自动递增 expect_no
            if parsed_expected and set(parsed_expected.keys()) == {0}:
                expected = [
                    ExpectedResult(expect_no=next_expect_no, description=parsed_expected[0])
                ]
            else:
                expected = [
                    ExpectedResult(expect_no=en, description=ed)
                    for en, ed in sorted(parsed_expected.items())
                    if ed
                ]

            if existing is not None:
                existing.steps.extend(steps)
                existing.expected.extend(expected)
                # 以首次出现为准；若后续行补充了空字段，则不覆盖
                if not existing.title and title:
                    existing.title = title
                if not existing.module and module:
                    existing.module = module
                if not existing.test_type and test_type:
                    existing.test_type = test_type
                if not existing.precondition and precondition:
                    existing.precondition = precondition
                if not existing.priority and priority:
                    existing.priority = priority
                if not existing.tags and tags:
                    existing.tags = tags
            else:
                cases_map[case_id] = ExcelCase(
                    case_id=case_id,
                    title=title,
                    steps=steps,
                    expected=expected,
                    module=module,
                    test_type=test_type,
                    precondition=precondition,
                    priority=priority,
                    tags=tags,
                )
    finally:
        close_fn()

    for case in cases_map.values():
        case.steps.sort(key=lambda s: s.step_no)

    return list(cases_map.values())


# ---------------------------------------------------------------------------
# 过滤与分组
# ---------------------------------------------------------------------------

def filter_cases_by_tags(cases: List[ExcelCase], tags: Optional[List[str]]) -> List[ExcelCase]:
    """按 tags 过滤用例"""
    if not tags:
        return cases
    tag_set = set(t.strip().lower() for t in tags if t)
    return [
        c for c in cases
        if tag_set
        & set(
            t.strip().lower()
            for t in _TAG_SPLIT_PATTERN.split(c.tags or "")
            if t.strip()
        )
    ]


def filter_cases_by_module(cases: List[ExcelCase], module: str) -> List[ExcelCase]:
    """按模块过滤"""
    if not module:
        return cases
    return [c for c in cases if c.module.strip().lower() == module.strip().lower()]


def filter_cases_by_priority(cases: List[ExcelCase], priorities: List[str]) -> List[ExcelCase]:
    """按优先级过滤，如 ["高", "中"]"""
    if not priorities:
        return cases
    pset = set(p.strip().lower() for p in priorities)
    return [c for c in cases if c.priority.strip().lower() in pset]


def group_cases_by_module(cases: List[ExcelCase]) -> dict[str, List[ExcelCase]]:
    """按模块分组"""
    groups: dict[str, List[ExcelCase]] = {}
    for c in cases:
        key = c.module or "未分类"
        groups.setdefault(key, []).append(c)
    return groups
