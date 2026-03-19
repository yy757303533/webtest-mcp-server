"""
webtest-mcp-server - MCP 入口

工具:
- list_projects_tool:  列出已配置项目
- get_grouped_cases:   按模块分组返回用例概要（用于分批规划）
- get_excel_cases:     读取 Excel 自然语言用例（支持按模块/标签/优先级过滤）
- generate_cases:      将 AI 分析的用例写入 Excel（需求文档 → Excel 全自动）
- save_test_results:   保存执行结果，生成 result.json / report.md / report.html
                       支持截图嵌入、按模块分批调用时累计合并
"""

import json
import os
from datetime import datetime, timezone
from pathlib import Path
from uuid import uuid4
from typing import Any, Optional

from mcp.server.fastmcp import FastMCP

from .config import get_projects_dir, list_projects, load_project_config
from .loader import (
    ExcelCase,
    filter_cases_by_module,
    filter_cases_by_priority,
    filter_cases_by_tags,
    group_cases_by_module,
    load_excel_cases,
)

try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

mcp = FastMCP("webtest_mcp_server")


def _is_relative_to(path: Path, base: Path) -> bool:
    """Py3.10 兼容的 is_relative_to"""
    try:
        path.relative_to(base)
        return True
    except ValueError:
        return False


def _get_excel_path(project_key: str, excel_path: str) -> Path:
    """解析 Excel 路径：相对则基于项目目录"""
    p = Path(excel_path)
    if p.is_absolute():
        return p.resolve()
    config = load_project_config(project_key)
    project_dir = Path(config.get("_project_dir", ".")).resolve()
    resolved = (project_dir / excel_path).resolve()
    if not _is_relative_to(resolved, project_dir):
        raise ValueError(f"非法 excel_path（越界项目目录）: {excel_path}")
    return resolved


def _get_artifacts_root(project: Optional[str] = None) -> Path:
    """
    结果落盘根目录优先级：
    1) WEBTEST_ARTIFACTS_DIR 环境变量（绝对/相对均可）
    2) 指定 project 时，落到 projects/<project>/artifacts（更贴近用例与配置）
    3) 兜底：当前工作目录下 artifacts（保持兼容）
    """
    env_dir = os.environ.get("WEBTEST_ARTIFACTS_DIR", "").strip()
    if env_dir:
        return Path(env_dir).expanduser().resolve()
    if project:
        try:
            config = load_project_config(project)
            project_dir = Path(config.get("_project_dir", ".")).resolve()
            return (project_dir / "artifacts").resolve()
        except Exception:
            # 若配置异常则继续走兜底
            pass
    return (Path.cwd() / "artifacts").resolve()


def _case_to_dict(c: ExcelCase) -> dict:
    """ExcelCase → 供 AI 消费的 dict（steps 和 expected 独立）"""
    d: dict[str, Any] = {
        "case_id": c.case_id,
        "title": c.title,
        "steps": [
            {"step_no": s.step_no, "description": s.description}
            for s in c.steps
        ],
        "expected": [
            {"expect_no": e.expect_no, "description": e.description}
            for e in c.expected
        ],
    }
    if c.module:
        d["module"] = c.module
    if c.precondition:
        d["precondition"] = c.precondition
    if c.priority:
        d["priority"] = c.priority
    if c.tags:
        d["tags"] = c.tags
    if c.test_type:
        d["test_type"] = c.test_type
    return d


def _load_and_filter(
    project: str,
    excel_path: str,
    tags: Optional[list[str]] = None,
    module: Optional[str] = None,
    priorities: Optional[list[str]] = None,
) -> tuple[list[ExcelCase], str]:
    """加载 + 过滤，返回 (cases, base_url)"""
    full_path = _get_excel_path(project, excel_path)
    if not full_path.exists():
        raise FileNotFoundError(f"Excel 不存在: {full_path}")
    cases = load_excel_cases(full_path)
    if tags:
        cases = filter_cases_by_tags(cases, tags)
    if module:
        cases = filter_cases_by_module(cases, module)
    if priorities:
        cases = filter_cases_by_priority(cases, priorities)
    config = load_project_config(project)
    base_url = config.get("base_url", "")
    return cases, base_url


# ───────────────────────────────────────────────────────────────────────────
# MCP 工具
# ───────────────────────────────────────────────────────────────────────────


@mcp.tool()
async def list_projects_tool() -> str:
    """
    列出所有已配置的 Web 测试项目
    Returns: JSON 格式的项目列表（含 projects、projects_dir）
    """
    projects = list_projects()
    projects_dir = get_projects_dir()
    return json.dumps(
        {"success": True, "projects": projects, "projects_dir": str(projects_dir)},
        ensure_ascii=False,
        indent=2,
    )


@mcp.tool()
async def get_excel_cases(
    project: str,
    excel_path: str,
    tags: Optional[list[str]] = None,
    module: Optional[str] = None,
    priorities: Optional[list[str]] = None,
) -> str:
    """
    读取 Excel 中的自然语言测试用例，供 AI 配合 playwright-mcp 执行。
    支持 .xls 和 .xlsx 格式。

    Args:
        project: 项目 key
        excel_path: Excel 文件路径（相对项目目录或绝对路径）
        tags: 可选，按标签过滤，如 ["导航", "CTA"]
        module: 可选，按模块过滤，如 "SPE模拟器"
        priorities: 可选，按优先级过滤，如 ["高", "中"]
    Returns:
        用例列表 JSON，每条含 case_id, title, module, precondition,
        steps: [{step_no, description, expected}], tags, priority
    """
    try:
        if project not in list_projects():
            return json.dumps(
                {"success": False, "error": f"项目 {project} 不存在"},
                ensure_ascii=False,
                indent=2,
            )
        cases, base_url = _load_and_filter(project, excel_path, tags, module, priorities)
        out = [_case_to_dict(c) for c in cases]
        resolved_path = str(_get_excel_path(project, excel_path))
        return json.dumps(
            {
                "success": True,
                "project": project,
                "base_url": base_url,
                "excel_path": str(excel_path),
                "excel_path_resolved": resolved_path,
                "count": len(out),
                "cases": out,
            },
            ensure_ascii=False,
            indent=2,
        )
    except Exception as e:
        return json.dumps(
            {"success": False, "error": str(e)},
            ensure_ascii=False,
            indent=2,
        )


@mcp.tool()
async def get_grouped_cases(
    project: str,
    excel_path: str,
    tags: Optional[list[str]] = None,
    priorities: Optional[list[str]] = None,
) -> str:
    """
    按模块分组返回用例概要，用于规划分批执行。
    不返回完整步骤，只返回每个模块的用例数量和 case_id 列表。

    AI 可据此决定先执行哪个模块，再调用 get_excel_cases(module=...) 获取详细步骤。

    Args:
        project: 项目 key
        excel_path: Excel 文件路径
        tags: 可选，按标签过滤
        priorities: 可选，按优先级过滤
    Returns:
        按模块分组的用例概要 JSON
    """
    try:
        if project not in list_projects():
            return json.dumps(
                {"success": False, "error": f"项目 {project} 不存在"},
                ensure_ascii=False,
                indent=2,
            )
        cases, base_url = _load_and_filter(
            project, excel_path, tags=tags, priorities=priorities
        )
        groups = group_cases_by_module(cases)
        resolved_path = str(_get_excel_path(project, excel_path))
        out = []
        for mod, mod_cases in groups.items():
            out.append(
                {
                    "module": mod,
                    "count": len(mod_cases),
                    "case_ids": [c.case_id for c in mod_cases],
                    "priorities": list(set(c.priority for c in mod_cases if c.priority)),
                }
            )
        # 按用例数降序
        out.sort(key=lambda g: g["count"], reverse=True)
        return json.dumps(
            {
                "success": True,
                "project": project,
                "base_url": base_url,
                "excel_path": str(excel_path),
                "excel_path_resolved": resolved_path,
                "total_cases": len(cases),
                "total_modules": len(out),
                "groups": out,
            },
            ensure_ascii=False,
            indent=2,
        )
    except Exception as e:
        return json.dumps(
            {"success": False, "error": str(e)},
            ensure_ascii=False,
            indent=2,
        )


@mcp.tool()
async def save_test_results(
    project: str,
    results: list[dict[str, Any]],
    excel_path: Optional[str] = None,
    screenshots: Optional[list[dict[str, str]]] = None,
) -> str:
    """
    保存测试执行结果，生成 result.json / report.md / report.html。

    每次调用都写入独立的 run_dir，同时将本次结果**追加**到项目级累计报告。
    按模块分批调用时，累计报告会持续合并所有已完成模块的结果。

    Args:
        project:     项目 key
        results:     结果列表，每项结构：
                       {
                         "case_id":  "HUUDI-T0001",
                         "title":    "First Name-空值拒绝",
                         "module":   "用户注册和登录",   # 可选，用于按模块汇总
                         "passed":   true,              # true=PASS / false=FAIL / null=SKIP
                         "error":    "E2 失败: xxx"     # 可选，FAIL/SKIP 时填写
                       }
        excel_path:  可选，用例文件名，用于追溯
        screenshots: 可选，失败截图列表，每项：
                       {
                         "case_id":   "HUUDI-T0002",    # 对应用例 ID
                         "path":      "/abs/path/to/screenshot.png"  # 绝对路径
                       }
                     截图会被复制到 run_dir/screenshots/ 并嵌入 HTML 报告。

    Returns:
        JSON，含 run_id、run_dir、report_html_path、latest_report_html_path、summary
    """
    import shutil
    import base64

    try:
        if project not in list_projects():
            return json.dumps(
                {"success": False, "error": f"项目 {project} 不存在"},
                ensure_ascii=False,
                indent=2,
            )

        artifacts_root = _get_artifacts_root(project)
        artifacts_dir = artifacts_root / project
        artifacts_dir.mkdir(parents=True, exist_ok=True)

        run_id = f"{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}-{uuid4().hex[:8]}"
        run_dir = artifacts_dir / run_id
        run_dir.mkdir(parents=True, exist_ok=True)

        # ── 处理截图 ──────────────────────────────────────────────────────
        # 将截图复制到 run_dir/screenshots/，key=case_id → base64 data URI
        screenshots_dir = run_dir / "screenshots"
        screenshot_map: dict[str, str] = {}   # case_id → data URI (for HTML embed)
        screenshot_paths: dict[str, str] = {} # case_id → relative path (for JSON)

        if screenshots:
            screenshots_dir.mkdir(exist_ok=True)
            for s in screenshots:
                cid = str(s.get("case_id", "")).strip()
                src = str(s.get("path", "")).strip()
                if not cid or not src:
                    continue
                src_path = Path(src)
                if not src_path.exists():
                    continue
                dest_name = f"{cid}{src_path.suffix or '.png'}"
                dest_path = screenshots_dir / dest_name
                shutil.copy2(src_path, dest_path)
                screenshot_paths[cid] = f"screenshots/{dest_name}"
                # Embed as base64 for self-contained HTML
                try:
                    b64 = base64.b64encode(dest_path.read_bytes()).decode()
                    ext = src_path.suffix.lstrip(".").lower() or "png"
                    mime = {"jpg": "jpeg", "jpeg": "jpeg", "png": "png", "webp": "webp"}.get(ext, "png")
                    screenshot_map[cid] = f"data:image/{mime};base64,{b64}"
                except Exception:
                    pass

        # ── 统计 ─────────────────────────────────────────────────────────
        passed  = sum(1 for r in results if r.get("passed") is True)
        failed  = sum(1 for r in results if r.get("passed") is False)
        skipped = sum(1 for r in results if r.get("passed") is None)
        timestamp = datetime.now(timezone.utc).isoformat()

        payload = {
            "timestamp": timestamp,
            "project": project,
            "excel_path": excel_path,
            "run_id": run_id,
            "summary": {
                "total": len(results),
                "passed": passed,
                "failed": failed,
                "skipped": skipped,
            },
            "screenshot_paths": screenshot_paths,
            "results": results,
        }
        result_path = run_dir / "result.json"
        with open(result_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

        # ── Markdown report ───────────────────────────────────────────────
        md_lines = [
            "# 测试结果",
            "",
            f"**项目**: {project}",
            f"**时间**: {timestamp}",
            f"**用例文件**: {excel_path or '-'}",
            f"**Run ID**: {run_id}",
            "",
            f"**总计**: {len(results)} 条，通过 {passed}，失败 {failed}，跳过 {skipped}",
            "",
            "| 用例ID | 模块 | 标题 | 结果 | 备注 |",
            "|--------|------|------|------|------|",
        ]
        for r in results:
            p = r.get("passed")
            status = "✅ PASS" if p is True else ("❌ FAIL" if p is False else "⏭️ SKIP")
            err   = str(r.get("error", "")).replace("|", "\\|")
            title = str(r.get("title", "")).replace("|", "\\|")
            mod   = str(r.get("module", "")).replace("|", "\\|")
            md_lines.append(f"| {r.get('case_id', '')} | {mod} | {title} | {status} | {err} |")
        report_md_path = run_dir / "report.md"
        report_md_path.write_text("\n".join(md_lines), encoding="utf-8")

        # ── HTML report (self-contained, screenshots embedded) ────────────
        def _html_escape(s: str) -> str:
            return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")

        pass_pct = round(passed / len(results) * 100) if results else 0
        rows_html = ""
        for r in results:
            p = r.get("passed")
            if p is True:
                badge, row_cls = '<span class="badge pass">PASS</span>', ""
            elif p is False:
                badge, row_cls = '<span class="badge fail">FAIL</span>', ' class="fail-row"'
            else:
                badge, row_cls = '<span class="badge skip">SKIP</span>', ' class="skip-row"'

            cid        = _html_escape(str(r.get("case_id", "")))
            mod        = _html_escape(str(r.get("module", "")))
            title_h    = _html_escape(str(r.get("title", "")))
            err_h      = _html_escape(str(r.get("error", "")))
            ss_tag     = ""
            if r.get("case_id") in screenshot_map:
                uri = screenshot_map[r["case_id"]]
                ss_tag = (
                    f'<div class="ss-thumb" onclick="openImg(this)">'
                    f'<img src="{uri}" alt="screenshot"></div>'
                )

            rows_html += (
                f'<tr{row_cls} data-status="{"pass" if p is True else ("fail" if p is False else "skip")}">'
                f'<td class="cid">{cid}</td>'
                f'<td class="mod">{mod}</td>'
                f'<td>{title_h}</td>'
                f'<td>{badge}</td>'
                f'<td class="err">{err_h}{ss_tag}</td>'
                f'</tr>\n'
            )

        html = f"""<!DOCTYPE html>
<html lang="zh">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{_html_escape(project)} 测试报告</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f5f5f7;color:#1d1d1f;padding:24px}}
h1{{font-size:22px;font-weight:600;margin-bottom:4px}}
.meta{{color:#6e6e73;font-size:13px;margin-bottom:20px;word-break:break-all}}
.cards{{display:flex;gap:12px;margin-bottom:20px;flex-wrap:wrap}}
.card{{background:#fff;border-radius:10px;padding:16px 20px;min-width:100px;box-shadow:0 1px 3px rgba(0,0,0,.08)}}
.card .num{{font-size:28px;font-weight:700;line-height:1}}
.card .lbl{{font-size:12px;color:#6e6e73;margin-top:4px}}
.card.pass .num{{color:#28a745}}.card.fail .num{{color:#dc3545}}
.card.skip .num{{color:#fd7e14}}.card.total .num{{color:#0066cc}}
.progress{{background:#e9ecef;border-radius:99px;height:8px;margin-bottom:20px;overflow:hidden}}
.progress-bar{{height:100%;border-radius:99px;background:#28a745}}
table{{width:100%;border-collapse:collapse;background:#fff;border-radius:10px;overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,.08)}}
th{{background:#f5f5f7;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.5px;padding:10px 12px;text-align:left;border-bottom:1px solid #e5e5ea}}
td{{padding:9px 12px;font-size:13px;border-bottom:1px solid #f0f0f0;vertical-align:top}}
tr:last-child td{{border-bottom:none}}
.fail-row td{{background:#fff8f8}}.skip-row td{{background:#fffbf5}}
.cid{{font-family:monospace;font-size:12px;color:#6e6e73;white-space:nowrap}}
.mod{{font-size:12px;color:#6e6e73;white-space:nowrap}}
.err{{color:#dc3545;font-size:12px;max-width:300px;word-break:break-word}}
.badge{{display:inline-block;padding:2px 8px;border-radius:99px;font-size:11px;font-weight:600;letter-spacing:.3px;white-space:nowrap}}
.badge.pass{{background:#d4edda;color:#155724}}.badge.fail{{background:#f8d7da;color:#721c24}}.badge.skip{{background:#fff3cd;color:#856404}}
.filter-bar{{display:flex;gap:8px;margin-bottom:16px;flex-wrap:wrap;align-items:center}}
.filter-btn{{padding:5px 14px;border:1px solid #d2d2d7;border-radius:99px;font-size:13px;cursor:pointer;background:#fff;transition:all .15s}}
.filter-btn.active{{background:#0066cc;color:#fff;border-color:#0066cc}}
.search{{padding:6px 14px;border:1px solid #d2d2d7;border-radius:99px;font-size:13px;outline:none;flex:1;min-width:160px}}
.ss-thumb{{margin-top:6px;cursor:zoom-in}}
.ss-thumb img{{max-width:120px;max-height:80px;border-radius:4px;border:1px solid #e5e5ea;display:block}}
.lightbox{{display:none;position:fixed;inset:0;background:rgba(0,0,0,.75);z-index:9999;align-items:center;justify-content:center}}
.lightbox.open{{display:flex}}
.lightbox img{{max-width:90vw;max-height:90vh;border-radius:8px;box-shadow:0 8px 32px rgba(0,0,0,.4)}}
</style>
</head>
<body>
<h1>{_html_escape(project)} 测试报告</h1>
<div class="meta">{_html_escape(timestamp)} &nbsp;|&nbsp; {_html_escape(excel_path or '-')} &nbsp;|&nbsp; Run: {run_id}</div>

<div class="cards">
  <div class="card total"><div class="num">{len(results)}</div><div class="lbl">总计</div></div>
  <div class="card pass"><div class="num">{passed}</div><div class="lbl">通过</div></div>
  <div class="card fail"><div class="num">{failed}</div><div class="lbl">失败</div></div>
  <div class="card skip"><div class="num">{skipped}</div><div class="lbl">跳过</div></div>
</div>

<div class="progress"><div class="progress-bar" style="width:{pass_pct}%"></div></div>

<div class="filter-bar">
  <button class="filter-btn active" onclick="setFilter('all',this)">全部 {len(results)}</button>
  <button class="filter-btn" onclick="setFilter('pass',this)">通过 {passed}</button>
  <button class="filter-btn" onclick="setFilter('fail',this)">失败 {failed}</button>
  <button class="filter-btn" onclick="setFilter('skip',this)">跳过 {skipped}</button>
  <input class="search" id="search" type="text" placeholder="搜索用例ID / 模块 / 标题..." oninput="doSearch()">
</div>

<table>
<thead><tr><th>用例ID</th><th>模块</th><th>标题</th><th>结果</th><th>备注 / 截图</th></tr></thead>
<tbody id="tbody">
{rows_html}
</tbody>
</table>

<div class="lightbox" id="lb" onclick="this.classList.remove('open')">
  <img id="lb-img" src="" alt="screenshot">
</div>

<script>
var tbody = document.getElementById('tbody');
var currentFilter = 'all';
function setFilter(type, btn) {{
  currentFilter = type;
  document.querySelectorAll('.filter-btn').forEach(function(b){{ b.classList.remove('active'); }});
  btn.classList.add('active');
  applyFilters();
}}
function doSearch() {{ applyFilters(); }}
function applyFilters() {{
  var q = document.getElementById('search').value.toLowerCase();
  Array.from(tbody.rows).forEach(function(r) {{
    var statusMatch = currentFilter === 'all' || r.dataset.status === currentFilter;
    var textMatch = !q || r.textContent.toLowerCase().includes(q);
    r.style.display = (statusMatch && textMatch) ? '' : 'none';
  }});
}}
function openImg(el) {{
  var img = el.querySelector('img');
  if (!img) return;
  document.getElementById('lb-img').src = img.src;
  document.getElementById('lb').classList.add('open');
}}
</script>
</body>
</html>"""

        html_report_path = run_dir / "report.html"
        html_report_path.write_text(html, encoding="utf-8")

        # ── 累计报告：合并历史 + 本次到 artifacts/<project>/ ─────────────
        # 读取已有累计结果（前几个模块的数据），追加本次结果
        cumulative_json_path = artifacts_dir / "result.json"
        cumulative_results: list[dict] = []
        if cumulative_json_path.exists():
            try:
                prev = json.loads(cumulative_json_path.read_text(encoding="utf-8"))
                cumulative_results = prev.get("results", [])
            except Exception:
                cumulative_results = []
        # 用 case_id 去重（同一 case 重跑时以最新为准）
        existing_ids = {r.get("case_id") for r in results}
        cumulative_results = [r for r in cumulative_results if r.get("case_id") not in existing_ids]
        cumulative_results.extend(results)

        cum_passed  = sum(1 for r in cumulative_results if r.get("passed") is True)
        cum_failed  = sum(1 for r in cumulative_results if r.get("passed") is False)
        cum_skipped = sum(1 for r in cumulative_results if r.get("passed") is None)
        cumulative_payload = {
            "timestamp": timestamp,
            "project": project,
            "excel_path": excel_path,
            "run_id": run_id,
            "summary": {
                "total": len(cumulative_results),
                "passed": cum_passed,
                "failed": cum_failed,
                "skipped": cum_skipped,
            },
            "results": cumulative_results,
        }
        cumulative_json_path.write_text(
            json.dumps(cumulative_payload, ensure_ascii=False, indent=2), encoding="utf-8"
        )

        # 累计 HTML 报告（复用同样的生成逻辑，只需重新渲染 rows）
        cum_pass_pct = round(cum_passed / len(cumulative_results) * 100) if cumulative_results else 0
        cum_rows_html = ""
        for r in cumulative_results:
            p = r.get("passed")
            if p is True:
                badge2, row_cls2 = '<span class="badge pass">PASS</span>', ""
            elif p is False:
                badge2, row_cls2 = '<span class="badge fail">FAIL</span>', ' class="fail-row"'
            else:
                badge2, row_cls2 = '<span class="badge skip">SKIP</span>', ' class="skip-row"'
            cid2   = _html_escape(str(r.get("case_id", "")))
            mod2   = _html_escape(str(r.get("module", "")))
            title2 = _html_escape(str(r.get("title", "")))
            err2   = _html_escape(str(r.get("error", "")))
            ss2 = ""
            if r.get("case_id") in screenshot_map:
                uri2 = screenshot_map[r["case_id"]]
                ss2 = f'<div class="ss-thumb" onclick="openImg(this)"><img src="{uri2}" alt="screenshot"></div>'
            cum_rows_html += (
                f'<tr{row_cls2} data-status="{"pass" if p is True else ("fail" if p is False else "skip")}">'
                f'<td class="cid">{cid2}</td><td class="mod">{mod2}</td>'
                f'<td>{title2}</td><td>{badge2}</td><td class="err">{err2}{ss2}</td></tr>\n'
            )

        cum_html = html.replace(
            f'<div class="num">{len(results)}</div><div class="lbl">总计</div>',
            f'<div class="num">{len(cumulative_results)}</div><div class="lbl">总计</div>',
        ).replace(
            f'<div class="num">{passed}</div><div class="lbl">通过</div>',
            f'<div class="num">{cum_passed}</div><div class="lbl">通过</div>',
        ).replace(
            f'<div class="num">{failed}</div><div class="lbl">失败</div>',
            f'<div class="num">{cum_failed}</div><div class="lbl">失败</div>',
        ).replace(
            f'<div class="num">{skipped}</div><div class="lbl">跳过</div>',
            f'<div class="num">{cum_skipped}</div><div class="lbl">跳过</div>',
        ).replace(
            f'style="width:{pass_pct}%"',
            f'style="width:{cum_pass_pct}%"',
        ).replace(
            f'>全部 {len(results)}<',
            f'>全部 {len(cumulative_results)}<',
        ).replace(
            f'>通过 {passed}<',
            f'>通过 {cum_passed}<',
        ).replace(
            f'>失败 {failed}<',
            f'>失败 {cum_failed}<',
        ).replace(
            f'>跳过 {skipped}<',
            f'>跳过 {cum_skipped}<',
        ).replace(
            f'\n{rows_html}\n',
            f'\n{cum_rows_html}\n',
        )

        latest_html_path = artifacts_dir / "report.html"
        latest_html_path.write_text(cum_html, encoding="utf-8")
        latest_md_path = artifacts_dir / "report.md"
        latest_md_path.write_text(report_md_path.read_text(encoding="utf-8"), encoding="utf-8")

        return json.dumps(
            {
                "success": True,
                "run_id": run_id,
                "run_dir": str(run_dir),
                "result_path": str(result_path),
                "report_html_path": str(html_report_path),
                "report_md_path": str(report_md_path),
                "latest_report_html_path": str(latest_html_path),
                "screenshots_copied": len(screenshot_map),
                "summary": payload["summary"],
                "cumulative_summary": cumulative_payload["summary"],
            },
            ensure_ascii=False,
            indent=2,
        )
    except Exception as e:
        return json.dumps(
            {"success": False, "error": str(e)},
            ensure_ascii=False,
            indent=2,
        )


@mcp.tool()
async def generate_cases(
    project: str,
    cases: list[dict[str, Any]],
    output_filename: str = "",
) -> str:
    """
    将 AI 从需求文档分析出的测试用例写入 Excel 文件，存放在项目目录。
    写入后可立即用 get_grouped_cases / get_excel_cases 读取并执行。

    Args:
        project: 项目 key（必须已在 projects/ 目录存在）
        cases: 用例列表，每项结构：
            {
              "case_id": "HUUDI-T0001",        # 必填，唯一标识
              "module": "用户注册和登录",         # 必填，功能模块
              "test_type": "功能测试",            # 可选
              "title": "First Name-空值拒绝",    # 必填，用例名称
              "precondition": "P1. 用户在注册页面", # 可选
              "steps": "S1. 留空\\nS2. 点击提交",  # 必填，测试步骤
              "expected": "E1. 显示错误提示",      # 必填，预期结果
              "priority": "高",                   # 可选：高/中/低
              "tags": "输入校验,必填"              # 可选，逗号分隔
            }
        output_filename: 输出文件名，默认自动生成带时间戳的名称
                         如 "huudi_cases_v2.xlsx"

    Returns:
        JSON，含写入路径和用例统计，可直接传给 get_excel_cases
    """
    if not _HAS_OPENPYXL:
        return json.dumps(
            {"success": False, "error": "generate_cases 需要 openpyxl，请执行: pip install openpyxl"},
            ensure_ascii=False, indent=2,
        )
    try:
        if project not in list_projects():
            return json.dumps(
                {"success": False, "error": f"项目 {project} 不存在"},
                ensure_ascii=False, indent=2,
            )

        config = load_project_config(project)
        project_dir = Path(config["_project_dir"]).resolve()

        # 自动生成文件名
        if not output_filename:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"cases_{ts}.xlsx"
        if not output_filename.endswith((".xlsx", ".xls")):
            output_filename += ".xlsx"

        output_path = (project_dir / output_filename).resolve()
        # 安全检查：不允许写到项目目录外
        try:
            output_path.relative_to(project_dir)
        except ValueError:
            return json.dumps(
                {"success": False, "error": f"output_filename 不能包含路径穿越: {output_filename}"},
                ensure_ascii=False, indent=2,
            )

        # 构建 Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TestCases"

        headers = [
            "用例编号", "模块", "测试类型", "用例说明（名称）",
            "预置条件", "测试步骤", "预期结果", "等级\n（高、中、低）", "场景标签",
        ]
        # 表头样式
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font_white = Font(bold=True, color="FFFFFF")
        wrap = Alignment(wrap_text=True, vertical="top")

        ws.append(headers)
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = wrap

        # 写入用例
        skipped = 0
        written = 0
        for c in cases:
            case_id = str(c.get("case_id", "")).strip()
            title = str(c.get("title", "")).strip()
            steps = str(c.get("steps", "")).strip()
            expected = str(c.get("expected", "")).strip()

            # 必填字段校验
            if not case_id or not title or not steps or not expected:
                skipped += 1
                continue

            row_data = [
                case_id,
                str(c.get("module", "")).strip(),
                str(c.get("test_type", "功能测试")).strip(),
                title,
                str(c.get("precondition", "")).strip(),
                steps,
                expected,
                str(c.get("priority", "中")).strip(),
                str(c.get("tags", "")).strip(),
            ]
            ws.append(row_data)
            # 自动换行
            for col_idx, _ in enumerate(row_data, 1):
                ws.cell(row=ws.max_row, column=col_idx).alignment = wrap
            written += 1

        # 列宽
        col_widths = [16, 18, 12, 30, 30, 40, 40, 10, 20]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

        wb.save(output_path)

        return json.dumps(
            {
                "success": True,
                "project": project,
                "output_file": output_filename,
                "output_path": str(output_path),
                "written": written,
                "skipped": skipped,
                "message": (
                    f"已写入 {written} 条用例到 {output_filename}。"
                    f"{'跳过 ' + str(skipped) + ' 条（缺少必填字段）。' if skipped else ''}"
                    f"现在可以调用 get_grouped_cases(project='{project}', excel_path='{output_filename}') 查看分组概要。"
                ),
            },
            ensure_ascii=False, indent=2,
        )
    except Exception as e:
        return json.dumps(
            {"success": False, "error": str(e)},
            ensure_ascii=False, indent=2,
        )


def main() -> None:
    """MCP 入口"""
    mcp.run()


if __name__ == "__main__":
    main()
